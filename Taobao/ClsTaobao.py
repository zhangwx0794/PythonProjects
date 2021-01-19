# -*- coding: utf-8 -*-
# @Time : 2021/1/18 0018 10:38
# @Author : Owen
# @Email : zhangwx0794@gmail.com
# @File : ClsTaobao.py
# @Project : Taobao

import xlwt, xlrd
import pymysql
import os
import re
import time
from openpyxl import *
import win32com.client as win32

class Taobao():

    # 1. 数据库操作returnCnt默认返回结果集条数，默认返回所有
    def sql_operation(self,sql,returnCnt=-1):
        conn = pymysql.Connect(
            host='47.111.113.238',
            port=3318,
            db='taobao',
            user='taobao',
            passwd='Boss123456..taobao',
            charset='utf8'
        )
        # 获取游标
        cursor = conn.cursor()
        try:
            cursor.execute(sql)
        except Exception as e:
            conn.rollback()
            print('异常sql: ',sql)
            print('事务处理失败!异常信息:', e)
        else:
            conn.commit()
        sqlRes = cursor.fetchall()
        # 关闭连接
        cursor.close()
        conn.close()
        if returnCnt == -1:
            return sqlRes
        elif returnCnt == 1:
            return sqlRes[0][0]
        else:
            return 0

    # 2. 查找当前工作目录下所有的xls文件
    def get_path_xls(self,absPath):
        allXls = os.listdir(absPath)
        zz = re.compile('(\.xls)$')
        xlsList = []
        for xls in allXls:
            zzRes = zz.findall(xls)
            if 'xls' in xls and len(zzRes) > 0:
                xlsList.append(xls)
        return xlsList

    # 3. 查找当前工作目录下所有的xlsx文件
    def get_path_xlsx(self,absPath):
        allXlsx = os.listdir(absPath)
        zz = re.compile('(\.xlsx)$')
        xlsxList = []
        for xlsx in allXlsx:
            zzRes = zz.findall(xlsx)
            if 'xlsx' in xlsx and len(zzRes) > 0:
                xlsxList.append(xlsx)
        return xlsxList

    # 4. 数据规范校验

    # 5. xls转xlsx
    def xls_to_xlsx(self,xlsPath):
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(xlsPath)
            wb.SaveAs(xlsPath + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            wb.Close()  # FileFormat = 56 is for .xls extension
            excel.Application.Quit()
        except Exception as e:
            print('xlsz转换xlsx异常',e)
        else:
            os.remove(xlsPath)
            print(xlsPath,'转换成功，源文件已删除')

    # 6. excel文件重命名
    def format_xls_name(self,xlsPath):
        # 1.拼接变量
        # * xls名称
        xlsName = str(xlsPath).split('\\')[-1]
        # * xls所在路径
        xlsPwd = str(xlsPath).split(xlsName)[0]
        # * 订单日期
        length1 = len(re.compile('^\d+\.\d+').findall(xlsName))
        if length1 > 0:
            rq = re.compile('^\d+\.\d+').findall(xlsName)[0]
            month = str(rq).split('.')[0]
            day = str(rq).split('.')[1]
            if int(month) >= 10:
                year = 2020
            else:
                year = 2021
            date = str(year) + '-' + str(month).rjust(2, '0') + '-' + str(day).rjust(2, '0')
            print(xlsName,date)
        else:
            length2 = len(re.compile('^\d{4}-\d{2}-\d{2}').findall(xlsName))
            if length2 > 0:
                date = re.compile('^\d{4}-\d{2}-\d{2}').findall(xlsName)[0]
            else:
                date = '2099-12-31'

        # * 店铺名称
        shopNameTemp = re.compile('[\u4e00-\u9fff]+').findall(xlsName)[0]
        shopName = str(shopNameTemp).replace('汇总', '').replace('订单', '').replace('总汇', '').replace('副本', '')
        # 2.拼接新的文件名称
        newXlsName = date + shopName + '.xlsx'
        newXlsPath = xlsPwd + newXlsName
        # * 判断新文件是否已经存在
        for i in range(2, 100):
            if os.path.exists(newXlsPath):
                print(newXlsPath, '新文件已存在，跳过重命名……','源文件名称：',xlsName)
                newXlsPath = xlsPwd + date + shopName + '_' + str(i) + '.xlsx'
            else:
                os.rename(xlsPath, newXlsPath)
                print('重命名成功!', xlsPath, ' => ', newXlsPath)
                break
        if '副本' in xlsName:
            print(xlsName, newXlsPath)
        return None

    # 7. 删除excel含有关键字的列
    def del_col_from_key(self,xlsPath, key_word,col=1):
        # xlsPath必须得是绝对路径
        wbb = load_workbook(xlsPath)
        wss = wbb.active
        # 删除第一列【时间】数据
        kw = str(wss.cell(1, col).value).strip()
        if str(key_word).strip() == kw:
            wss.delete_cols(col)
            wbb.save(xlsPath)
            print(xlsPath + '第一列关键字' + key_word + '删除成功!')
        wbb.close()
        return None

    # 8. 校验指定范围列数据有效性
    def list_none_check(self,row, startCol, endCol):
        for i in range(startCol, endCol):
            if str(row[i]) == '':
                # print('none',end=' ')
                return 0
        return 1

    # 9. 订单号唯一检测
    # 查询数据库检查订单号是否已存在
    def chk_data_is_exist(self,order_id):
        sql = 'select count(0) from orderInfo where orderId = %s' % order_id
        cnt = self.sql_operation(sql,1)
        # 返回订单号重复查询结果 存在返回1 不存在返回0
        return cnt

    # 9.1 批量检测订单号是否唯一
    def chkXlsOrderUniq(self,xlsPath):
        # 将数据库中所有的订单号存入数组orderIdLst中
        sql = 'select orderId from orderInfo where isDel = 0'
        orderIdRes = self.sql_operation(sql,-1)
        orderIdLst = []
        for orderId in orderIdRes:
            orderIdLst.append(str(orderId[0]).strip())
        # 获取xls中所有订单号
        wb = xlrd.open_workbook(xlsPath)
        # * 打开第一个sheet
        ws = wb.sheet_by_index(0)
        cnt = 0
        for line in range(1,ws.nrows):
            colValue = ws.cell_value(rowx=line,colx=5)
            if colValue in orderIdLst:
                # print(xlsPath,'订单号与数据库重复',colValue,'第',line,'行')
                cnt += 1
        return cnt

    # 10. 数据导入
    def importData(self,xlsPath):
        xlsName = str(xlsPath).split('\\')[-1]
        wb = xlrd.open_workbook(xlsPath)
        # * 打开第一个sheet
        ws = wb.sheet_by_index(0)
        # * 从第二行开始导入数据
        if self.data_format_check(xlsPath) == 0 and self.chkXlsOrderUniq(xlsPath) == 0:
            dataImportNum = 0
            for line in range(1, ws.nrows):
                # * 获取当前行数据
                rowList = ws.row_values(line)
                # * 校验当前行指定范围列数据是否完整
                # 日期、经手人、店铺名称、宝贝名称、关键词、旺旺ID、订单号、客单价、佣金
                sqlFormat = 'insert into orderInfo(shopName,goodsName,goodsKey,wangwangId,orderId,goodsPrice,goodsYj,redPackets,ssyj,handlerName,opWechatId,custName,date) values({0},{1},{2},{3},{4},{5},{6},{7},{8},{9},{10},{11},{12})'
                shopName = '\'' + str(rowList[1]) + '\''
                goodsName = '\'' + str(rowList[2]) + '\''
                goodsKey = '\'' + str(rowList[3]) + '\''
                wangwangId = '\'' + str(rowList[4]).strip() + '\''
                orderId = '\'' + str(rowList[5]).strip() + '\''
                goodsPrice = rowList[6]
                goodsYj = rowList[7]
                redPackets = rowList[8]
                ssyj = rowList[9]
                handlerName = '\'' + str(rowList[10]) + '\''
                opWechatId = '\'' + str(rowList[11]) + '\''
                custName = '\'' + str(rowList[12]) + '\''
                date = '\''+str(re.compile('^\d{4}-\d{2}-\d{2}').findall(xlsName)[0])+'\''
                sql = sqlFormat.format(shopName, goodsName, goodsKey, wangwangId, orderId, goodsPrice, goodsYj,
                                       redPackets, ssyj, handlerName, opWechatId, custName, date)
                # 根据订单号检查数据，如果不重复，则将表格中的数据插入数据库；0不存在  1存在
                try:
                    # print('插入数据库……', sql)
                    self.sql_operation(sql)
                    dataImportNum += 1
                except Exception as e:
                    print(xlsName, '有毛病，插入数据库异常')
                    return 0
            print(xlsName, '成功导入{0}条数据'.format(dataImportNum))
            return dataImportNum
    # 11. 店铺名&旺旺ID唯一检测

    # 12. 更新店铺名

    # 13. 删除重复的xls文件
    def delRepeName(self,xlsPath,xlsxList):
        xlsName = str(xlsPath).split('\\')[-1]
        if xlsName+'x' in xlsxList:
            os.remove(xlsPath)
            print('已删除', xlsPath)

    # 14. 取出文件中第一列的值
    def getColValues(self,xlsPath,col=1):
        wbb = load_workbook(xlsPath)
        wss = wbb.active
        # 删除第一列【时间】数据
        kw = str(wss.cell(1, col).value).strip()
        wbb.close()
        return kw
    # 15. 写入数据到excel
    def writeData2Xls(self,xlsPath,value,row=1,col=1):
        try:
            wbb = load_workbook(xlsPath)
            wss = wbb.active
            # 写入数据内容到单元格中
            wss.cell(row,col).value = value
            wbb.save(xlsPath)
            wbb.close()
        except Exception as e:
            print('写入数据异常',e)
            return 0
        else:
            return 1

    # 16. 插入新列
    def insertColum(self,xlsPath,col):
        try:
            wbb = load_workbook(xlsPath)
            wss = wbb.active
            # 插入新列
            wss.insert_cols(idx=col)
            wbb.save(xlsPath)
            wbb.close()
        except Exception as e:
            print('插入列异常',e)
            return 0
        else:
            return 1

    # 17. 删除空行
    def delteBlankRow(self,xlsPath):
        try:
            wbb = load_workbook(xlsPath)
            wss = wbb.active
            # 写入数据内容到单元格中
            for row in range(1,wss.max_row+1):
                rowList = []
                for col in range(1,wss.max_column+1):
                    if wss.cell(row=row,column=col).value != None:
                        rowList.append(wss.cell(row=row,column=col).value)
                if len(rowList) <= 1 or 'SUM' in str(rowList):
                    print(xlsPath,'第',row,'行为空，已删除',rowList)
                    wss.delete_rows(idx=row)
                    wbb.save(xlsPath)
                    wbb.close()
                    return -1
                else:
                    pass
                    # print(xlsPath,rowList)
            wbb.save(xlsPath)
            wbb.close()
        except Exception as e:
            print('删除列异常',e)
            return -1
        else:
            return 0

    # 18. 根据规则表格完善
    def completeForm(self,xlsPath,col,value):
        try:
            wbb = load_workbook(xlsPath)
            wss = wbb.active
            # 写入数据内容到单元格中，从第2行开始填充数据
            for row in range(2,wss.max_row+1):
                # 红包及其他
                if wss.cell(row=row,column=col).value == None and col == 9:
                    wss.cell(row=row, column=col).value = 0
                # 店铺名称
                elif wss.cell(row=row,column=col).value == None and col == 2:
                    wss.cell(row=row, column=col).value = value
                # 刷手佣金
                elif wss.cell(row=row,column=col).value == None and col == 10:
                    wss.cell(row=row, column=col).value = 0
                # 经手人
                elif wss.cell(row=row,column=col).value == None and col == 11:
                    wss.cell(row=row, column=col).value = value
                # 操作人微信
                elif wss.cell(row=row,column=col).value == None and col == 12:
                    wss.cell(row=row, column=col).value = value
                # 客户名称
                elif col == 13:
                    wss.cell(row=row, column=col).value = value
                # 日期
                elif col == 14:
                    wss.cell(row=row, column=col).value = value
            wbb.save(xlsPath)
            wbb.close()
        except Exception as e:
            print('插入列异常',e)
            return 0
        else:
            return 1
    # 19. 删除订单号为空或订单号重复的行
    def delBlankOrderRow(self,xlsPath):
        while (True):
            # 接收错误行号
            line = self.chkRepeOrderInXls(xlsPath)
            if line > 0:
                try:
                    wbb = load_workbook(xlsPath)
                    wss = wbb.active
                    # 删除错误行
                    wss.delete_rows(idx=line)
                    print('delete',xlsPath,'第',line,'行')
                    wbb.save(xlsPath)
                except Exception as e:
                    print('删除列异常',e)
                    wbb.close()
                finally:
                    wbb.close()
            else:
                break
    # 20. 数据规范检查
    def data_format_check(self,xlsPath):
        wb = xlrd.open_workbook(xlsPath)
        # * 打开第一个sheet
        ws = wb.sheet_by_index(0)
        for line in range(1, ws.nrows):
            # * 获取当前行数据
            rowList = list(ws.row_values(line))
            if rowList[5] == '':
                print(xlsPath, '第', line + 1, '行', '第', 5, '列数据为空!')
                return -1
            for i in range(5, 10):
                if not str(rowList[i]).replace('.', '').isdigit():
                    # print('错误: 第',i+1,'列数据不规范，含有非数字或小数点字符!')
                    return -2
        return 0
    # 21. 检测单张表是否有重复订单号，正常返回0，异常返回订单号所在行号
    def chkRepeOrderInXls(self,xlsPath):
        try:
            wb = xlrd.open_workbook(xlsPath)
            ws = wb.sheet_by_index(0)
            lstTmp = []
            for line in range(1, ws.nrows):
                if ws.cell(line,5).value not in lstTmp and ws.cell(line,5).value != '':
                    lstTmp.append(ws.cell(line,5).value)
                elif ws.cell(line,5).value == '':
                    print(xlsPath,'订单号为空，第',line+1,'行')
                    return line+1
                else:
                    print(xlsPath,'有订单号重复,第',line+1,'行')
                    return line+1
        except Exception as e:
            print(xlsPath,'捕捉到打开表异常',e)
            return -1
        return 0
    # 99. 测试方法
    def testFunc(self):
        print('\n'+os.getcwd())
        pass

